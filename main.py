# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import tkinter as tk
import openpyxl
import pathlib
import random
from apscheduler.schedulers.background import BackgroundScheduler
import os
from datetime import datetime
from slugify import slugify

answers_saved_path = "C:\\Answers"


class Timer:
    def __init__(self):
        self.minute = 0
        self.seconds = 5

    def reset(self):
        self.minute = 0
        self.seconds = 5

    def current_timer(self):
        return str(self.minute) + ":" + str(self.seconds)

    def update_timer(self):
        if self.seconds == 0:
            if self.minute == 0:
                pass
            else:
                self.minute -= 1
                self.seconds = 59
        else:
            self.seconds -= 1


def get_random_card():
    clear_input_text()
    if input_text.cget("state") == "disabled":
        input_text.config(state="normal", background="white")
    label_topic.config(text=reader.get_random_topic())
    timer.reset()


def save_answers():
    ans = input_text.get("1.0", 'end-1c')
    file_location = input_save_location.get("1.0", "end-1c")

    if not os.path.isdir(file_location):
        label_file_saved.config(text="'{}' is not valid directory".format(file_location), fg="red")
        return

    filename = slugify(reader.current_topic) + datetime.now().strftime("%d-%b-%Y-%H-%M-%S") + ".txt"

    full_file_path = os.path.join(file_location, filename)
    try:
        file = open(full_file_path, "w+")
        file.write(ans)
        file.close()
        label_file_saved.config(text="File saved successfully", fg="green")
    except Exception as e:
        label_file_saved.config(text="There was an error saving a file", fg="red")
        print("cannot write to file {}".format(e))


def clear_input_text():
    if input_text.cget("state") == "normal":
        input_text.delete('1.0', tk.END)
    else:
        input_text.config(state="normal")
        input_text.delete('1.0', tk.END)
        input_text.config(state="disabled")


def update_stopwatch():
    timer.update_timer()
    label_timer.config(text=timer.current_timer())
    if timer.seconds == 0 and timer.minute == 0:
        input_text.config(state="disabled", background="light grey")

    # save file


class BPTReader:
    def __init__(self):
        self.topics = []
        self.workbook = None
        self.parse_file()
        self.current_topic = None

    def get_max_rows(self):
        return self.workbook.active.max_row

    def parse_file(self):
        current_path = pathlib.Path().resolve()
        settings_file = open(str(current_path) + "\\settings.txt", "r")
        excel_file_path = settings_file.readline()
        settings_file.close()
        self.workbook = openpyxl.load_workbook(excel_file_path)

    def get_all_topics(self):
        for row in range(1, self.get_max_rows() + 1):
            topic = self.workbook.active.cell(row=row, column=2)
            if topic.value is not None:
                self.topics.append(topic.value)
        return self.topics

    def get_random_topic(self):
        self.current_topic = random.choice(self.topics)
        return self.current_topic


if __name__ == "__main__":
    reader = BPTReader()
    print(reader.get_all_topics())

    timer = Timer()

    window = tk.Tk()
    window.title("Flash card revision")
    window.geometry('400x400')

    label_topic = tk.Label(window, text=reader.get_random_topic(), fg="red", font="Arial 12 bold")
    label_topic.place(x=80, y=40)

    label_timer = tk.Label(window, text=timer.current_timer(), fg="black")
    label_timer.place(x=80, y=70)

    button = tk.Button(window, text="Get card", fg="black", command=get_random_card)
    button.place(x=80, y=100)

    button_clear = tk.Button(window, text="Clear", fg="black", command=clear_input_text)
    button_clear.place(x=80, y=400)

    button_save = tk.Button(window, text="Save", fg="black", command=save_answers)
    button_save.place(x=125, y=400)

    input_text = tk.Text(window, height=15, width=100)
    input_text.place(x=80, y=150)

    label_save_location = tk.Label(window, text="location:")
    label_save_location.place(x=140, y=70)

    input_save_location = tk.Text(window, height=1, width=80)
    input_save_location.insert(tk.INSERT, answers_saved_path)
    input_save_location.place(x=195, y=70)

    schedule = BackgroundScheduler()
    schedule.add_job(update_stopwatch, "interval", seconds=1)
    schedule.start()

    label_file_saved = tk.Label(window)
    label_file_saved.place(x=80, y=430)

    window.mainloop()
